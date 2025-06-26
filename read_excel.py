import openpyxl

workbook = openpyxl.load_workbook('/home/ubuntu/upload/Análise.xlsx')

# Get sheet names
sheet_names = workbook.sheetnames
print(f'Nomes das planilhas: {sheet_names}')

# Read 'Metas' sheet
metas_sheet = workbook['Metas']
print('\nConteúdo da planilha Metas:')
for row in metas_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Alça' sheet
alca_sheet = workbook['Alça ']
print('\nConteúdo da planilha Alça:')
for row in alca_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Fundo' sheet
fundo_sheet = workbook['Fundo ']
print('\nConteúdo da planilha Fundo:')
for row in fundo_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Topo' sheet
topo_sheet = workbook['Topo ']
print('\nConteúdo da planilha Topo:')
for row in topo_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Acessório' sheet
acessorio_sheet = workbook['Acessório ']
print('\nConteúdo da planilha Acessório:')
for row in acessorio_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Líner' sheet
liner_sheet = workbook['Líner ']
print('\nConteúdo da planilha Líner:')
for row in liner_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Mesa' sheet
mesa_sheet = workbook['Mesa ']
print('\nConteúdo da planilha Mesa:')
for row in mesa_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Corte-Carimbadeira' sheet
corte_carimbadeira_sheet = workbook['Corte-Carimbadeira']
print('\nConteúdo da planilha Corte-Carimbadeira:')
for row in corte_carimbadeira_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Geral' sheet
geral_sheet = workbook['Geral']
print('\nConteúdo da planilha Geral:')
for row in geral_sheet.iter_rows(values_only=True):
    print(row)

# Read 'fechamento' sheet
fechamento_sheet = workbook['fechamento']
print('\nConteúdo da planilha fechamento:')
for row in fechamento_sheet.iter_rows(values_only=True):
    print(row)

# Read 'Planilha2' sheet
planilha2_sheet = workbook['Planilha2']
print('\nConteúdo da planilha Planilha2:')
for row in planilha2_sheet.iter_rows(values_only=True):
    print(row)


